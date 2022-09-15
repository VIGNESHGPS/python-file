import openpyxl
class hotel:
    def hotelname(self):
        namefile=openpyxl.load_workbook("C:\\Users\\GOD\\OneDrive\\Documents\\VGCN BOOK CAL.xlsx")#to open the excel work book
        sheet=namefile["appu"]# to open the respective sheet
        cell=sheet.cell(row=1,column=6).value
        print(cell)
        sheet.cell(row=1 , column=2).value="dharshini mental" # to write the value on excel sheet
        namefile.save("C:\\Users\\GOD\\OneDrive\\Documents\\BOOK CAL.xlsx")

        #to find maxi number
        print("...............****************................")
        maximumrow = sheet.max_row
        print(maximumrow)
        maximumcolumn = sheet.max_column
        print(maximumcolumn)

        print("...............****************................")
        write_new_xlfile=openpyxl.Workbook()
        sheetname=write_new_xlfile.active

        #to create new excel sheet and write value on it(copyiedc value from above accessed xl sheet)
        for rowvalue in range(1, maximumrow + 1):  # row
            for columnvalue in range(1, maximumcolumn + 1):  # column
                print(sheet.cell(row=rowvalue, column=columnvalue).value)
                eachvalue = sheet.cell(row=rowvalue, column=columnvalue).value
                sheetname.cell(row=rowvalue, column=columnvalue).value = eachvalue
                # sheetoutputname.cell(row=rowvalue, column=columnvalue).value = "entred through script"
        write_new_xlfile.save("C:\\Users\\GOD\\OneDrive\\Documents\\BOOK CAL 2.xlsx")

    #enter the value in the excel sheet through the list
    def name_list(self):
        list=["vignesh","preethi","hari priya","gokul","pushpa dharshini","hari prasadh","samyuktha","seenu tharan","gunalini"]
        new_file=openpyxl.Workbook()#CREATING BOOK
        new_sheet=new_file.active


        for row_value in range(1,len(list)):
            a=list[row_value]
            print(a)
            new_sheet.cell(row=row_value,column=1).value=a
        new_file.save("C:\\Users\\GOD\\OneDrive\\Documents\\namelist.xlsx")
        new_file.close()
obj=hotel()
obj.hotelname()
obj.name_list()